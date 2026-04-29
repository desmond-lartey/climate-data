"""
============================================================
WEST AFRICA PRECIPITATION PRODUCTS — DECISION SUPPORT TOOL
============================================================
Generates WA_Precipitation_Decision_Tool.xlsx automatically
by reading validation CSV outputs from DATA_DIR.

Run from the precipitation_assessment folder:
    python generate_decision_tool.py

Required input files (all produced by the Python pipeline):
    DATA_DIR/validation_by_zone.csv
    DATA_DIR/validation_overall.csv
    DATA_DIR/product_ranking_by_zone.csv
    DATA_DIR/threshold_sensitivity.csv

Output:
    outputs/WA_Precipitation_Decision_Tool.xlsx
============================================================
"""

import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl as xl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
import sys

# ════════════════════════════════════════════════════════════
# § 0  PATHS — edit BASE_DIR if needed
# ════════════════════════════════════════════════════════════

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "DATA_DIR"
OUT_DIR  = BASE_DIR / "outputs"
OUT_DIR.mkdir(exist_ok=True)
OUT_FILE = OUT_DIR / "WA_Precipitation_Decision_Tool.xlsx"

REQUIRED_FILES = {
    "validation_by_zone"   : DATA_DIR / "validation_by_zone.csv",
    "validation_overall"   : DATA_DIR / "validation_overall.csv",
    "product_ranking_zone" : DATA_DIR / "product_ranking_by_zone.csv",
    "threshold_sensitivity": DATA_DIR / "threshold_sensitivity.csv",
}

# ════════════════════════════════════════════════════════════
# § 1  LOAD AND VALIDATE DATA
# ════════════════════════════════════════════════════════════

def load_data():
    """Load all required CSVs and check they exist."""
    missing = [k for k, p in REQUIRED_FILES.items() if not p.exists()]
    if missing:
        print("\nERROR — missing required files:")
        for k in missing:
            print(f"  {REQUIRED_FILES[k]}")
        print("\nRun the pipeline first:")
        print("  python validation_metrics.py")
        print("  python threshold_sensitivity.py")
        sys.exit(1)

    print("Loading data files...")
    data = {k: pd.read_csv(p) for k, p in REQUIRED_FILES.items()}

    # Add All West Africa rows to validation_by_zone using overall data
    vo = data["validation_overall"]
    extra = []
    for _, row in vo.iterrows():
        r = {"zone_name": "All West Africa", "product": row["product"],
             "n_stations": 16, "n_station_months": 3840}
        for col in ["n","bias","pbias","mae","rmse","r","r2","nse","kge",
                    "pod","far","csi","ets","freq_bias",
                    "hits","misses","false_al","correct_neg"]:
            r[col] = row.get(col, np.nan)
        extra.append(r)
    data["vbz_full"] = pd.concat(
        [data["validation_by_zone"], pd.DataFrame(extra)],
        ignore_index=True
    )

    # Summary
    for k, df in data.items():
        if k != "vbz_full":
            print(f"  {k}: {df.shape[0]} rows × {df.shape[1]} cols")

    zones    = sorted(data["vbz_full"]["zone_name"].unique())
    products = sorted(data["vbz_full"]["product"].unique())
    print(f"\n  Zones    : {zones}")
    print(f"  Products : {products}")
    return data


# ════════════════════════════════════════════════════════════
# § 2  SCORING ENGINE
# ════════════════════════════════════════════════════════════

APP_WEIGHTS = {
    "Fire risk monitoring":    {"kge":0.10,"r":0.20,"nse":0.05,"pod":0.15,
                                "far":0.30,"csi":0.15,"pbias":0.05},
    "Wildlife & habitat":      {"kge":0.25,"r":0.20,"nse":0.15,"pod":0.15,
                                "far":0.10,"csi":0.10,"pbias":0.05},
    "Drought early warning":   {"kge":0.15,"r":0.10,"nse":0.10,"pod":0.30,
                                "far":0.20,"csi":0.10,"pbias":0.05},
    "Flood forecasting":       {"kge":0.10,"r":0.10,"nse":0.10,"pod":0.20,
                                "far":0.10,"csi":0.25,"pbias":0.15},
    "Agricultural planning":   {"kge":0.20,"r":0.25,"nse":0.15,"pod":0.15,
                                "far":0.10,"csi":0.10,"pbias":0.05},
    "Hydrological modelling":  {"kge":0.30,"r":0.20,"nse":0.25,"pod":0.10,
                                "far":0.05,"csi":0.05,"pbias":0.05},
    "Climate trend analysis":  {"kge":0.15,"r":0.25,"nse":0.25,"pod":0.10,
                                "far":0.10,"csi":0.10,"pbias":0.05},
}

APP_FOCUS = {
    "Fire risk monitoring":   "FAR critical — false rain forecast = unpreparedness",
    "Wildlife & habitat":     "KGE + r — habitat depends on magnitude and seasonal pattern",
    "Drought early warning":  "POD critical — missed dry spell = missed intervention",
    "Flood forecasting":      "CSI + POD — capturing extreme events matters most",
    "Agricultural planning":  "r + KGE — seasonal onset and totals equally important",
    "Hydrological modelling": "KGE + NSE — standard water balance benchmarks",
    "Climate trend analysis":  "r + NSE — temporal consistency over absolute accuracy",
}

ZONE_NOTES = {
    "Saharian":         "Hyperarid. FAR unstable at any threshold. Use MERRA-2 or PERSIANN-CDR. Never rely on absolute PBIAS alone.",
    "Sahelian":         "Unimodal wet season Jul-Sep. GPM-IMERG leads on KGE. Report threshold sensitivity in publications.",
    "Soudanian":        "Reliable rainfall. CHIRPS and GPM-IMERG consistently strong. Safe for hydrological modelling.",
    "Guinean":          "Bimodal Jun+Oct. Most products overestimate magnitude. Prioritise r over KGE. Avoid TerraClimate for absolutes.",
    "Guineo-Congolean": "High-rainfall equatorial. GPM-IMERG best KGE. TerraClimate severely overestimates — avoid for water balance.",
    "All West Africa":  "Pooled across 16 stations and 5 zones. GPM-IMERG leads overall. Zone-specific selection always preferred.",
}


def weighted_score(row, weights):
    """Compute weighted composite score from a data row."""
    def norm(v, mn=0, mx=1, inv=False):
        if pd.isna(v): return 0.0
        n = min(1.0, max(0.0, (float(v)-mn)/(mx-mn)))
        return 1.0-n if inv else n

    return round(
        weights["kge"]   * norm(row.get("kge",   np.nan), -1,  1)     +
        weights["r"]     * norm(row.get("r",      np.nan),  0,  1)     +
        weights["nse"]   * norm(row.get("nse",    np.nan), -5,  1)     +
        weights["pod"]   * norm(row.get("pod",    np.nan),  0,  1)     +
        weights["far"]   * norm(row.get("far",    np.nan),  0,  1, True) +
        weights["csi"]   * norm(row.get("csi",    np.nan),  0,  1)     +
        weights["pbias"] * norm(abs(row.get("pbias", 0) or 0), 0, 60, True),
    4)


def build_scores_table(vbz_full):
    """Pre-compute weighted scores for every app × zone × product."""
    rows = []
    zones    = sorted(vbz_full["zone_name"].unique())
    products = sorted(vbz_full["product"].unique())
    for app, weights in APP_WEIGHTS.items():
        for zone in zones:
            sub = vbz_full[vbz_full["zone_name"] == zone]
            for prod in products:
                prow = sub[sub["product"] == prod]
                if prow.empty: continue
                s = weighted_score(prow.iloc[0].to_dict(), weights)
                rows.append({"app": app, "zone": zone,
                             "product": prod, "score": s,
                             "score_key": app + zone + prod})
    return pd.DataFrame(rows)


# ════════════════════════════════════════════════════════════
# § 3  STYLE HELPERS
# ════════════════════════════════════════════════════════════

C = dict(
    dark="1C2D40", mid="2E4057", light="F4F6F9", white="FFFFFF",
    text="1A1A2E", muted="6B7280", green="1D9E75", blue="378ADD",
    amber="EF9F27", red="C0392B", border="D0D7E0",
)
ZONE_COL = {
    "Saharian":         "C8A96E",
    "Sahelian":         "E8A838",
    "Soudanian":        "CC6600",
    "Guinean":          "78C850",
    "Guineo-Congolean": "1A6B1A",
    "All West Africa":  "185FA5",
}
PROD_COL = {
    "CHIRPS":        "1f77b4",
    "ERA5_LAND":     "9467bd",
    "GPM_IMERG":     "d62728",
    "MERRA2":        "e377c2",
    "PERSIANN_CDR":  "ff7f0e",
    "TERRACLIMATE":  "17becf",
}


def F(bold=False, color=None, sz=10, italic=False):
    return Font(bold=bold, color=color or C["text"], size=sz,
                italic=italic, name="Arial")

def Fill(color):
    return PatternFill("solid", fgColor=color)

def Align(h="left", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical="center",
                     wrap_text=wrap, indent=indent)

def thin_side():
    return Side(style="thin", color=C["border"])

def ThinBorder():
    s = thin_side()
    return Border(left=s, right=s, top=s, bottom=s)

def AccentBorder(color):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def W(ws, row, col, val, bg=None, fg=None, bold=False, sz=10,
      align="left", wrap=False, italic=False, num_fmt=None):
    c = ws.cell(row, col, val)
    c.font   = F(bold, fg or C["text"], sz, italic)
    c.fill   = Fill(bg or C["white"])
    c.alignment = Align(align, wrap)
    if num_fmt: c.number_format = num_fmt
    return c

def SectionHdr(ws, row, c1, c2, text, bg=None, sz=11):
    ws.merge_cells(
        f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
    c = ws.cell(row, c1, text)
    c.font   = F(True, C["white"], sz)
    c.fill   = Fill(bg or C["mid"])
    c.alignment = Align("left", indent=1)
    ws.row_dimensions[row].height = 22
    return c

def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════
# § 4  BUILD WORKBOOK
# ════════════════════════════════════════════════════════════

def build_workbook(data, scores_df):
    wb = xl.Workbook()
    wb.remove(wb.active)

    vbz_full = data["vbz_full"]
    ts       = data["threshold_sensitivity"]
    prz      = data["product_ranking_zone"]

    zones    = ["Saharian","Sahelian","Soudanian",
                "Guinean","Guineo-Congolean","All West Africa"]
    products = sorted(vbz_full["product"].unique())
    apps     = list(APP_WEIGHTS.keys())

    # ── Find key column in vbz_full ─────────────────────────
    vbz_cols    = list(vbz_full.columns) + ["LOOKUP_KEY"]
    key_col_idx = len(vbz_cols)           # 1-based col in sheet
    key_col_ltr = get_column_letter(key_col_idx)

    # ── Threshold key col ─────────────────────────────────
    ts_cols      = list(ts.columns) + ["LOOKUP_KEY"]
    ts_key_idx   = len(ts_cols)
    ts_key_ltr   = get_column_letter(ts_key_idx)

    # ── Scores key col ────────────────────────────────────
    sc_cols     = ["app","zone","product","score","score_key"]
    sc_key_idx  = 5   # "score_key" is col 5
    sc_key_ltr  = get_column_letter(sc_key_idx)

    # ════════════════════════════════════════════════════
    # Sheet 1: DATA_by_zone
    # ════════════════════════════════════════════════════
    ws1 = wb.create_sheet("DATA_by_zone")
    ws1.freeze_panes = "A2"

    orig_cols = list(vbz_full.columns)
    all_cols  = orig_cols + ["LOOKUP_KEY"]

    for ci, h in enumerate(all_cols, 1):
        c = ws1.cell(1, ci, h)
        c.font   = F(True, C["white"], 9)
        c.fill   = Fill(C["dark"])
        c.alignment = Align("center")
        ws1.column_dimensions[get_column_letter(ci)].width = \
            max(9, len(h) + 2)

    for ri, (_, row) in enumerate(vbz_full.iterrows(), 2):
        bg   = C["light"] if ri % 2 == 0 else C["white"]
        zone = str(row.get("zone_name",""))
        for ci, col in enumerate(orig_cols, 1):
            v = row[col]
            if pd.isna(v): v = None
            c = ws1.cell(ri, ci, v)
            c.font  = F(sz=9,
                        color=C["white"] if ci == 1 else C["text"],
                        bold=(ci == 1))
            c.fill  = Fill(ZONE_COL.get(zone, bg) if ci == 1 else bg)
            c.alignment = Align("center" if ci > 2 else "left")
            if isinstance(v, float) and v is not None:
                c.number_format = "0.0000"
        # LOOKUP_KEY
        c = ws1.cell(ri, key_col_idx, f"=A{ri}&B{ri}")
        c.font  = F(sz=9, color=C["muted"])
        c.fill  = Fill(bg)
        c.alignment = Align("center")

    ws1.auto_filter.ref = f"A1:{get_column_letter(len(orig_cols))}1"
    ws1.sheet_properties.tabColor = "888888"

    # ════════════════════════════════════════════════════
    # Sheet 2: DATA_overall
    # ════════════════════════════════════════════════════
    ws2 = wb.create_sheet("DATA_overall")
    ws2.freeze_panes = "A2"
    vo = data["validation_overall"]
    for ci, h in enumerate(vo.columns, 1):
        c = ws2.cell(1, ci, h)
        c.font  = F(True, C["white"], 9)
        c.fill  = Fill(C["dark"])
        c.alignment = Align("center")
        ws2.column_dimensions[get_column_letter(ci)].width = \
            max(9, len(h) + 2)
    for ri, (_, row) in enumerate(vo.iterrows(), 2):
        bg   = C["light"] if ri % 2 == 0 else C["white"]
        prod = str(row.get("product",""))
        for ci, col in enumerate(vo.columns, 1):
            v = row[col]
            if pd.isna(v): v = None
            c = ws2.cell(ri, ci, v)
            c.font  = F(sz=9,
                        bold=(ci==1),
                        color=PROD_COL.get(prod, C["text"]) if ci==1
                              else C["text"])
            c.fill  = Fill(bg)
            c.alignment = Align("center" if ci > 1 else "left")
            if isinstance(v, float) and v is not None:
                c.number_format = "0.0000"
    ws2.sheet_properties.tabColor = "888888"

    # ════════════════════════════════════════════════════
    # Sheet 3: DATA_threshold
    # ════════════════════════════════════════════════════
    ws3 = wb.create_sheet("DATA_threshold")
    ws3.freeze_panes = "A2"
    ts_orig = list(ts.columns)
    for ci, h in enumerate(ts_cols, 1):
        c = ws3.cell(1, ci, h)
        c.font  = F(True, C["white"], 9)
        c.fill  = Fill(C["dark"])
        c.alignment = Align("center")
        ws3.column_dimensions[get_column_letter(ci)].width = \
            max(9, len(h) + 2)
    for ri, (_, row) in enumerate(ts.iterrows(), 2):
        bg = C["light"] if ri % 2 == 0 else C["white"]
        for ci, col in enumerate(ts_orig, 1):
            v = row[col]
            if pd.isna(v): v = None
            c = ws3.cell(ri, ci, v)
            c.font  = F(sz=9)
            c.fill  = Fill(bg)
            c.alignment = Align("center")
            if isinstance(v, float) and v is not None:
                c.number_format = "0.0000"
        # LOOKUP_KEY: threshold_col & zone_col & product_col
        # Find column positions dynamically
        t_idx = ts_orig.index("threshold") + 1 if "threshold" in ts_orig else 1
        z_idx = ts_orig.index("zone")      + 1 if "zone"      in ts_orig else 2
        p_idx = ts_orig.index("product")   + 1 if "product"   in ts_orig else 3
        t_ltr = get_column_letter(t_idx)
        z_ltr = get_column_letter(z_idx)
        p_ltr = get_column_letter(p_idx)
        c = ws3.cell(ri, ts_key_idx,
                     f'=TEXT({t_ltr}{ri},"0.0")&{z_ltr}{ri}&{p_ltr}{ri}')
        c.font  = F(sz=9, color=C["muted"])
        c.fill  = Fill(bg)
    ws3.sheet_properties.tabColor = "888888"

    # ════════════════════════════════════════════════════
    # Sheet 4: SCORES (pre-computed weighted scores)
    # ════════════════════════════════════════════════════
    ws4 = wb.create_sheet("SCORES")
    ws4.freeze_panes = "A2"
    for ci, h in enumerate(sc_cols, 1):
        c = ws4.cell(1, ci, h)
        c.font  = F(True, C["white"], 9)
        c.fill  = Fill(C["dark"])
        c.alignment = Align("center")
        ws4.column_dimensions[get_column_letter(ci)].width = \
            max(10, len(h) + 2)
    ws4.column_dimensions["E"].width = 35  # score_key

    for ri, (_, row) in enumerate(scores_df.iterrows(), 2):
        bg = C["light"] if ri % 2 == 0 else C["white"]
        for ci, col in enumerate(sc_cols, 1):
            v = row[col]
            c = ws4.cell(ri, ci, v)
            c.font  = F(sz=9)
            c.fill  = Fill(bg)
            c.alignment = Align(
                "center" if ci > 3 else "left")
            if col == "score":
                c.number_format = "0.0000"
    ws4.sheet_properties.tabColor = "888888"

    scores_end = len(scores_df) + 1  # last data row in SCORES sheet

    # ════════════════════════════════════════════════════
    # Sheet 5: APP_WEIGHTS
    # ════════════════════════════════════════════════════
    ws5 = wb.create_sheet("APP_WEIGHTS")
    ws5.freeze_panes = "A2"
    ws5.sheet_view.showGridLines = False
    aw_hdrs = ["Application","kge","r","nse","pod","far",
               "csi","pbias","Primary focus"]
    for ci, h in enumerate(aw_hdrs, 1):
        c = ws5.cell(1, ci, h)
        c.font  = F(True, C["white"], 9)
        c.fill  = Fill(C["dark"])
        c.alignment = Align("center")
        ws5.column_dimensions[get_column_letter(ci)].width = \
            max(10, len(h) + 2)
    ws5.column_dimensions["A"].width = 24
    ws5.column_dimensions["I"].width = 40

    for ri, (app, w) in enumerate(APP_WEIGHTS.items(), 2):
        bg = C["light"] if ri % 2 == 0 else C["white"]
        ws5.row_dimensions[ri].height = 20
        W(ws5, ri, 1, app, bg=bg, bold=True, sz=10)
        for ci, m in enumerate(["kge","r","nse","pod","far","csi","pbias"], 2):
            c = ws5.cell(ri, ci, w[m])
            c.font   = F(sz=10)
            c.fill   = Fill(bg)
            c.alignment = Align("center")
            c.number_format = "0%"
        c = ws5.cell(ri, 9, APP_FOCUS.get(app, ""))
        c.font  = F(sz=9, italic=True, color=C["muted"])
        c.fill  = Fill(bg)
        c.alignment = Align("left", wrap=True)
    ws5.sheet_properties.tabColor = C["amber"]

    # ════════════════════════════════════════════════════
    # Sheet 6: SELECTOR (main user tool)
    # ════════════════════════════════════════════════════
    ws6 = wb.create_sheet("SELECTOR")
    ws6.sheet_view.showGridLines = False
    ws6.sheet_view.zoomScale = 90

    set_col_widths(ws6, {
        "A":2,"B":24,"C":20,"D":11,"E":11,
        "F":10,"G":10,"H":10,"I":10,"J":10,"K":10,"L":10,"M":3,
    })

    # Title
    ws6.row_dimensions[1].height = 6
    ws6.merge_cells("B2:L2")
    c = ws6.cell(2, 2,
        "West Africa Precipitation Product Selector")
    c.font   = F(True, C["white"], 15)
    c.fill   = Fill(C["dark"])
    c.alignment = Align("center")
    ws6.row_dimensions[2].height = 34

    ws6.merge_cells("B3:L3")
    c = ws6.cell(3, 2,
        f"Auto-generated from validation data · "
        f"{len(products)} products · "
        f"{len([z for z in zones if z != 'All West Africa'])} zones · "
        f"2001–2020")
    c.font   = F(False, "AAAAAA", 10, italic=True)
    c.fill   = Fill(C["mid"])
    c.alignment = Align("center")
    ws6.row_dimensions[3].height = 18
    ws6.row_dimensions[4].height = 8

    # ── Step 1: Dropdowns ────────────────────────────────
    SectionHdr(ws6, 5, 2, 11,
               "STEP 1 — Select your management context", C["mid"])

    for ci, lbl in enumerate(
        ["Management application","Ecological zone",
         "Detection threshold","Uncertainty tolerance"], 2):
        W(ws6, 6, ci, lbl,
          bg=C["light"], fg=C["muted"], bold=True, sz=9, align="center")
    ws6.row_dimensions[6].height = 18

    defaults = [apps[0], "Sahelian", "1.0",
                "Low — need high confidence"]
    for ci, dv in enumerate(defaults, 2):
        c = ws6.cell(7, ci, dv)
        c.font   = F(True, C["dark"], 11)
        c.fill   = Fill(C["white"])
        c.alignment = Align("center")
        c.border = AccentBorder(C["green"])
    ws6.row_dimensions[7].height = 28

    app_list   = ",".join(apps)
    zone_list  = ",".join(zones)
    thresh_list = "0.1,0.5,1.0,2.0,5.0"
    unc_list   = ("Low — need high confidence,"
                  "Medium — ensemble acceptable,"
                  "High — exploratory use")

    for dv_range, formula in [
        ("B7", f'"{app_list}"'),
        ("C7", f'"{zone_list}"'),
        ("D7", f'"{thresh_list}"'),
        ("E7", f'"{unc_list}"'),
    ]:
        dv = DataValidation(type="list", formula1=formula,
                            showDropDown=False)
        ws6.add_data_validation(dv)
        dv.add(dv_range)

    ws6.row_dimensions[8].height = 8
    ws6.merge_cells("B9:L9")
    c = ws6.cell(9, 2,
        "Changing the application updates metric weights — "
        "rankings shift to reflect what matters for that use case. "
        "Zone controls which validation data is used.")
    c.font   = F(False, C["muted"], 9, italic=True)
    c.fill   = Fill(C["light"])
    c.alignment = Align("left", wrap=True, indent=1)
    ws6.row_dimensions[9].height = 20
    ws6.row_dimensions[10].height = 8

    # ── Weights display ──────────────────────────────────
    SectionHdr(ws6, 11, 2, 11,
               "Active metric weights for selected application",
               C["mid"])
    wt_metrics = ["KGE","r","NSE","POD","FAR","CSI","PBIAS"]
    for ci, m in enumerate(wt_metrics, 2):
        W(ws6, 12, ci, m,
          bg=C["dark"], fg=C["white"], bold=True, sz=9, align="center")
    ws6.row_dimensions[12].height = 18

    # Weight formulas pulling from APP_WEIGHTS sheet
    aw_metric_cols = {"KGE":"B","r":"C","NSE":"D",
                      "POD":"E","FAR":"F","CSI":"G","PBIAS":"H"}
    for ci, m in enumerate(wt_metrics, 2):
        col = aw_metric_cols[m]
        f   = (f'=IFERROR(INDEX(APP_WEIGHTS!${col}:${col},'
               f'MATCH($B$7,APP_WEIGHTS!$A:$A,0)),"")')
        c   = ws6.cell(13, ci, f)
        c.font   = F(True, C["green"], 11)
        c.fill   = Fill(C["light"])
        c.alignment = Align("center")
        c.number_format = "0%"
    ws6.row_dimensions[13].height = 22

    ws6.merge_cells("B14:L14")
    c = ws6.cell(14, 2,
        '=IFERROR(INDEX(APP_WEIGHTS!$I:$I,'
        'MATCH($B$7,APP_WEIGHTS!$A:$A,0)),"")')
    c.font   = F(False, C["amber"], 10, italic=True)
    c.fill   = Fill("FFFBF0")
    c.alignment = Align("left", wrap=True, indent=1)
    ws6.row_dimensions[14].height = 22
    ws6.row_dimensions[15].height = 8

    # ── Product ranking table ────────────────────────────
    SectionHdr(ws6, 16, 2, 11,
               "STEP 2 — Ranked products (updates with selection)",
               C["mid"])

    rank_hdrs = ["Rank","Product","Weighted\nscore",
                 "KGE","r","POD","FAR","CSI","NSE","PBIAS%"]
    for ci, h in enumerate(rank_hdrs, 2):
        c = ws6.cell(17, ci, h)
        c.font   = F(True, C["white"], 9)
        c.fill   = Fill(C["dark"])
        c.alignment = Align("center", wrap=True)
    ws6.row_dimensions[17].height = 28

    # Column map: metric name → column letter in DATA_by_zone
    metric_map = {
        "KGE":"M","r":"J","POD":"N","FAR":"O",
        "CSI":"P","NSE":"L","PBIAS%":"G",
    }

    for i, prod in enumerate(products):
        r  = 18 + i
        bg = C["light"] if i % 2 == 0 else C["white"]
        ws6.row_dimensions[r].height = 24

        # Rank — RANK against score column
        c = ws6.cell(r, 2,
            f'=IF(D{r}=0,"",RANK(D{r},$D$18:$D${17+len(products)},0))')
        c.font   = F(True, C["dark"], 13)
        c.fill   = Fill(bg)
        c.alignment = Align("center")

        # Product name
        c = ws6.cell(r, 3, prod)
        c.font   = F(True, PROD_COL.get(prod, C["text"]), 11)
        c.fill   = Fill(bg)
        c.alignment = Align("center")

        # Weighted score — MATCH on pre-computed SCORES sheet
        score_f = (f'=IFERROR(INDEX(SCORES!$D:$D,'
                   f'MATCH($B$7&$C$7&"{prod}",'
                   f'SCORES!${sc_key_ltr}:${sc_key_ltr},0)),0)')
        c = ws6.cell(r, 4, score_f)
        c.font   = F(True, C["green"], 11)
        c.fill   = Fill(bg)
        c.alignment = Align("center")
        c.number_format = "0.000"

        # Individual metrics from DATA_by_zone
        for ci, (mn, dl) in enumerate(metric_map.items(), 5):
            f = (f'=IFERROR(INDEX(DATA_by_zone!${dl}:${dl},'
                 f'MATCH($C$7&"{prod}",'
                 f'DATA_by_zone!${key_col_ltr}:${key_col_ltr},0)),"-")')
            c = ws6.cell(r, ci, f)
            c.font   = F(sz=10,
                         color=C["red"] if mn=="FAR" else C["text"])
            c.fill   = Fill(bg)
            c.alignment = Align("center")
            c.number_format = "0.000"

    ws6.row_dimensions[18 + len(products)].height = 8

    # ── Zone context ─────────────────────────────────────
    SectionHdr(ws6, 19+len(products), 2, 11,
               "STEP 3 — Zone context & cautions", C["mid"])

    for zi, zone in enumerate(zones):
        r    = 20 + len(products) + zi
        note = ZONE_NOTES.get(zone, "")
        zcol = ZONE_COL.get(zone, C["mid"])
        ws6.row_dimensions[r].height = 34
        c = ws6.cell(r, 2, zone)
        c.font   = F(True, C["white"], 10)
        c.fill   = Fill(zcol)
        c.alignment = Align("center")
        ws6.merge_cells(f"C{r}:L{r}")
        c2 = ws6.cell(r, 3, note)
        c2.font  = F(sz=10)
        c2.fill  = Fill(C["light"] if zi % 2 == 0 else C["white"])
        c2.alignment = Align("left", wrap=True, indent=1)

    row_after_zones = 20 + len(products) + len(zones) + 1
    ws6.row_dimensions[row_after_zones].height = 8

    # ── Threshold mini-table ─────────────────────────────
    SectionHdr(ws6, row_after_zones+1, 2, 11,
               "Detection metrics at selected threshold — all products",
               C["mid"])

    th_hdrs = ["Product","POD","FAR","CSI","ETS","Freq.bias"]
    for ci, h in enumerate(th_hdrs, 2):
        W(ws6, row_after_zones+2, ci, h,
          bg=C["dark"], fg=C["white"], bold=True, sz=9, align="center")
    ws6.row_dimensions[row_after_zones+2].height = 18

    # Column positions for threshold metrics in DATA_threshold
    # threshold=col1, zone=col2, product=col3, pod=col4, far=col5,
    # csi=col6, ets=col7, freq_bias=col8 — find dynamically
    ts_col_map = {col: i+1 for i, col in enumerate(ts.columns)}
    th_metric_cols = {
        "POD": get_column_letter(ts_col_map.get("pod", 4)),
        "FAR": get_column_letter(ts_col_map.get("far", 5)),
        "CSI": get_column_letter(ts_col_map.get("csi", 6)),
        "ETS": get_column_letter(ts_col_map.get("ets", 7)),
        "Freq.bias": get_column_letter(ts_col_map.get("freq_bias", 8)),
    }

    for i, prod in enumerate(products):
        r  = row_after_zones + 3 + i
        bg = C["light"] if i % 2 == 0 else C["white"]
        ws6.row_dimensions[r].height = 22
        c = ws6.cell(r, 2, prod)
        c.font   = F(True, PROD_COL.get(prod, C["text"]), 10)
        c.fill   = Fill(bg)
        c.alignment = Align("center")

        for ci, (mn, dl) in enumerate(th_metric_cols.items(), 3):
            f = (f'=IFERROR(INDEX(DATA_threshold!${dl}:${dl},'
                 f'MATCH(TEXT($D$7,"0.0")&$C$7&"{prod}",'
                 f'DATA_threshold!${ts_key_ltr}:${ts_key_ltr},0)),"-")')
            c = ws6.cell(r, ci, f)
            c.font   = F(sz=10,
                         color=C["red"] if mn=="FAR" else C["text"])
            c.fill   = Fill(bg)
            c.alignment = Align("center")
            c.number_format = "0.000"

    ws6.sheet_properties.tabColor = C["green"]

    # ════════════════════════════════════════════════════
    # Sheet 7: SCORECARD
    # ════════════════════════════════════════════════════
    ws7 = wb.create_sheet("SCORECARD")
    ws7.sheet_view.showGridLines = False
    set_col_widths(ws7, {"A":2,"B":30,"C":12,"D":10,"E":10,"F":36})

    ws7.merge_cells("B2:F2")
    c = ws7.cell(2, 2, "Full Metric Scorecard — by Zone and Product")
    c.font   = F(True, C["white"], 14)
    c.fill   = Fill(C["dark"])
    c.alignment = Align("center")
    ws7.row_dimensions[2].height = 30
    ws7.row_dimensions[1].height = 6

    for ci, lbl in enumerate(["Zone","Product"], 2):
        W(ws7, 3, ci, lbl,
          bg=C["light"], fg=C["muted"], bold=True, sz=9, align="center")
        default = "Sahelian" if lbl=="Zone" else products[0]
        c = ws7.cell(4, ci, default)
        c.font   = F(True, C["dark"], 11)
        c.fill   = Fill(C["white"])
        c.alignment = Align("center")
        c.border = AccentBorder(C["blue"])

    dv_z = DataValidation(type="list", formula1=f'"{zone_list}"')
    dv_p = DataValidation(type="list",
                          formula1=f'"{",".join(products)}"')
    ws7.add_data_validation(dv_z)
    ws7.add_data_validation(dv_p)
    dv_z.add("B4"); dv_p.add("C4")
    ws7.row_dimensions[3].height = 18
    ws7.row_dimensions[4].height = 26
    ws7.row_dimensions[5].height = 10

    # Metric rows — (is_hdr, label, data_col, perfect, interpretation)
    sc_rows = [
        (True,  "Continuous performance metrics","","",""),
        (False, "Metric","Value","Perfect","Interpretation"),
        (False, "KGE — Kling-Gupta efficiency","M","1.0",
         "Primary ranking metric. Decomposes into r (timing), α (variability), β (bias). <−0.41 = worse than mean."),
        (False, "r — Pearson correlation","J","1.0",
         "Seasonal pattern agreement. >0.85 = good. Key for phenology, fire onset, migration."),
        (False, "NSE — Nash-Sutcliffe efficiency","L","1.0",
         "Hydrological benchmark. <0 = worse than predicting observed mean."),
        (False, "Bias — mean error (mm/day)","F","0.0",
         "Positive = product wetter. Useful for water balance applications."),
        (False, "PBIAS — percent bias (%)","G","0%",
         ">±25% indicates significant systematic error. Comparable across zones."),
        (False, "MAE — mean absolute error","H","0.0",
         "Average error magnitude. Less sensitive to extremes than RMSE."),
        (False, "RMSE — root mean squared error","I","0.0",
         "Penalises peak errors heavily. Key for flood risk applications."),
        (True,  "Categorical detection metrics (1.0 mm/day threshold)","","",""),
        (False, "Metric","Value","Perfect","Interpretation"),
        (False, "POD — probability of detection","N","1.0",
         "Fraction of wet months detected. >0.85 = good. Critical for drought early warning."),
        (False, "FAR — false alarm ratio","O","0.0",
         "Fraction of wet predictions that were dry. <0.15 = acceptable for fire applications."),
        (False, "CSI — critical success index","P","1.0",
         "Combined detection skill. >0.75 = good for conservation planning."),
        (False, "ETS — equitable threat score","Q","1.0",
         "Detection adjusted for chance. >0.3 = meaningful skill. Robust in dry zones."),
        (False, "Freq. bias","R","1.0",
         ">1 = over-predicts rain frequency. <1 = under-predicts."),
    ]

    for i, (is_hdr, label, dcol, perfect, interp) in \
            enumerate(sc_rows):
        r  = 6 + i
        ws7.row_dimensions[r].height = 24 if not is_hdr else 22
        if is_hdr:
            ws7.merge_cells(f"B{r}:F{r}")
            c = ws7.cell(r, 2, label)
            c.font   = F(True, C["white"], 10)
            c.fill   = Fill(C["mid"])
            c.alignment = Align("left", indent=1)
        elif label == "Metric":
            for ci, txt in enumerate(
                    [label,"Value","Perfect","Interpretation"], 2):
                c = ws7.cell(r, ci, txt)
                c.font  = F(True, C["muted"], 9)
                c.fill  = Fill(C["light"])
                c.alignment = Align(
                    "left" if ci == 2 else "center")
            ws7.merge_cells(f"E{r}:F{r}")
        else:
            bg = C["light"] if i % 2 == 0 else C["white"]
            W(ws7, r, 2, label, bg=bg, sz=10)
            if dcol:
                f = (f'=IFERROR(INDEX(DATA_by_zone!${dcol}:${dcol},'
                     f'MATCH($B$4&$C$4,'
                     f'DATA_by_zone!${key_col_ltr}:${key_col_ltr},0)),"-")')
                c = ws7.cell(r, 3, f)
                c.font   = F(True, C["dark"], 10)
                c.fill   = Fill(bg)
                c.alignment = Align("center")
                c.number_format = "0.000"
            W(ws7, r, 4, perfect, bg=bg, fg=C["green"],
              sz=10, align="center")
            ws7.merge_cells(f"E{r}:F{r}")
            W(ws7, r, 5, interp, bg=bg, fg=C["muted"],
              sz=9, wrap=True)

    ws7.sheet_properties.tabColor = C["blue"]

    # ════════════════════════════════════════════════════
    # Sheet 8: README
    # ════════════════════════════════════════════════════
    ws8 = wb.create_sheet("README")
    ws8.sheet_view.showGridLines = False
    set_col_widths(ws8, {"A":2,"B":28,"C":60})
    ws8.row_dimensions[1].height = 6

    ws8.merge_cells("B2:C2")
    c = ws8.cell(2, 2,
        "West Africa Precipitation Decision Tool — README")
    c.font   = F(True, C["white"], 14)
    c.fill   = Fill(C["dark"])
    c.alignment = Align("center")
    ws8.row_dimensions[2].height = 30

    readme_rows = [
        ("HOW TO USE", "", True),
        ("SELECTOR sheet", "Main tool. Use the 4 dropdowns to select application, zone, threshold, and uncertainty tolerance. The product ranking and metric values update automatically.", False),
        ("SCORECARD sheet", "Select any zone and product to see the complete 13-metric profile with interpretation guidance.", False),
        ("DATA sheets", "DATA_by_zone, DATA_overall, DATA_threshold — raw validation outputs. Do not edit these manually.", False),
        ("SCORES sheet", "Pre-computed weighted scores for all app × zone × product combinations. Used by SELECTOR formulas.", False),
        ("APP_WEIGHTS sheet", "Metric weight profiles per application. Edit these to customise the scoring for your own priorities.", False),
        ("HOW TO REFRESH DATA", "", True),
        ("After new pipeline run", "Replace contents of DATA_by_zone, DATA_overall, DATA_threshold by running: python generate_decision_tool.py", False),
        ("Script reads", f"DATA_DIR/validation_by_zone.csv, validation_overall.csv, product_ranking_by_zone.csv, threshold_sensitivity.csv", False),
        ("ECOLOGICAL ZONES", "", True),
        ("Saharian", "Hyperarid (<25 mm/yr). 1 station: Nouakchott (WA016). FAR structurally elevated — use with caution.", False),
        ("Sahelian", "200-600 mm/yr. 5 stations: Dakar, Bamako, Ouagadougou, Niamey, Banjul.", False),
        ("Soudanian", "600-1200 mm/yr. 3 stations: Abuja, Conakry, Kano.", False),
        ("Guinean", "1200-2000 mm/yr. 3 stations: Freetown, Monrovia, Kumasi.", False),
        ("Guineo-Congolean", ">2000 mm/yr. 4 stations: Accra, Abidjan, Lomé, Cotonou.", False),
    ]

    for i, (label, desc, is_hdr) in enumerate(readme_rows):
        r  = 3 + i
        bg = C["mid"] if is_hdr else (C["light"] if i%2==0 else C["white"])
        ws8.row_dimensions[r].height = 28 if not is_hdr else 22
        c1 = ws8.cell(r, 2, label)
        c1.font  = F(True, C["white"] if is_hdr else C["dark"], 10)
        c1.fill  = Fill(bg)
        c1.alignment = Align("left", indent=1)
        c2 = ws8.cell(r, 3, desc)
        c2.font  = F(sz=10, color=C["text"])
        c2.fill  = Fill(bg)
        c2.alignment = Align("left", wrap=True, indent=1)

    ws8.sheet_properties.tabColor = C["amber"]

    # ── Sheet order ───────────────────────────────────────
    wb.move_sheet("README",     offset=-99)
    wb.move_sheet("SELECTOR",   offset=-98)
    wb.move_sheet("SCORECARD",  offset=-97)

    return wb


# ════════════════════════════════════════════════════════════
# § 5  MAIN
# ════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("\n" + "="*60)
    print("  WA Precipitation Decision Tool — Generator")
    print("="*60)
    print(f"  DATA_DIR : {DATA_DIR}")
    print(f"  Output   : {OUT_FILE}")

    data      = load_data()
    scores_df = build_scores_table(data["vbz_full"])
    print(f"\n  Scores table: {len(scores_df)} rows "
          f"({len(APP_WEIGHTS)} apps × "
          f"{data['vbz_full']['zone_name'].nunique()} zones × "
          f"{data['vbz_full']['product'].nunique()} products)")

    print("\nBuilding workbook...")
    wb = build_workbook(data, scores_df)

    wb.save(OUT_FILE)
    print(f"\n{'='*60}")
    print(f"  DONE — saved to: {OUT_FILE}")
    print(f"{'='*60}")
    print(f"  Sheets: {wb.sheetnames}")
    print("""
  To refresh after a new pipeline run:
      python generate_decision_tool.py
    """)
